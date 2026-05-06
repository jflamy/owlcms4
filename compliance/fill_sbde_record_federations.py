from openpyxl import load_workbook
from pathlib import Path

SOURCE = Path(r'c:\Dev\git\owlcms-meets\competitions\ccsr2026\preparation\SBDE_2026-05-05_08h14;20.xlsx')
TARGET = SOURCE.with_name('SBDE_2026-05-05_08h14;20_record-federations.xlsx')

QUEBEC_NAMES = {'quebec', 'québec'}
CANADIAN_PROVINCES = {
    'alberta',
    'british columbia',
    'manitoba',
    'new brunswick',
    'newfoundland and labrador',
    'nova scotia',
    'ontario',
    'prince edward island',
    'saskatchewan',
    'northwest territories',
    'nunavut',
    'yukon',
}
PANAM_COMMONWEALTH = {
    'antigua and barbuda',
    'bahamas',
    'barbados',
    'belize',
    'canada',
    'dominica',
    'grenada',
    'guyana',
    'jamaica',
    'saint kitts and nevis',
    'saint lucia',
    'saint vincent and the grenadines',
    'trinidad and tobago',
}
COMMONWEALTH_OTHER = {
    'australia',
    'new zealand',
    'solomon islands',
    'india',
    'england',
    'scotland',
    'wales',
    'northern ireland',
    'singapore',
    'malaysia',
    'nauru',
    'samoa',
    'tonga',
    'kiribati',
    'tuvalu',
    'vanuatu',
    'papua new guinea',
    'fiji',
    'cyprus',
    'malta',
    'brunei',
    'pakistan',
    'bangladesh',
    'sri lanka',
    'south africa',
    'botswana',
    'cameroon',
    'eswatini',
    'gabon',
    'gambia',
    'ghana',
    'kenya',
    'lesotho',
    'malawi',
    'mauritius',
    'mozambique',
    'namibia',
    'nigeria',
    'rwanda',
    'seychelles',
    'sierra leone',
    'tanzania',
    'uganda',
    'zambia',
}
PANAM_NON_COMMONWEALTH = {
    'argentina',
    'bolivia',
    'brazil',
    'chile',
    'colombia',
    'costa rica',
    'cuba',
    'dominican republic',
    'ecuador',
    'el salvador',
    'guatemala',
    'haiti',
    'honduras',
    'mexico',
    'nicaragua',
    'panama',
    'paraguay',
    'peru',
    'suriname',
    'united states',
    'uruguay',
    'venezuela',
}


def normalize(value):
    return str(value).strip().lower() if value is not None else ''


def federations_for_team(team_name):
    key = normalize(team_name)
    if key in QUEBEC_NAMES:
        return 'IWF,PanAm,CWF,WCH,FHQ'
    if key in CANADIAN_PROVINCES:
        return 'IWF,PanAm,CWF,WCH'
    if key in PANAM_COMMONWEALTH:
        return 'IWF,PanAm,CWF'
    if key in PANAM_NON_COMMONWEALTH:
        return 'IWF,PanAm'
    if key == 'australia':
        return 'IWF,CWF,AUS'
    if key in COMMONWEALTH_OTHER:
        return 'IWF,CWF'
    raise ValueError(f'No federation rule for team: {team_name!r}')


wb = load_workbook(SOURCE)
ws = wb['Athletes']
updated = []
for row in range(10, ws.max_row + 1):
    last_name = ws.cell(row=row, column=3).value
    if last_name in (None, ''):
        break
    team = ws.cell(row=row, column=5).value
    value = federations_for_team(team)
    ws.cell(row=row, column=16).value = value
    updated.append((row, last_name, team, value))

wb.save(TARGET)
print(f'WROTE\t{TARGET}')
print(f'UPDATED_ROWS\t{len(updated)}')
for sample in updated[:10]:
    print('SAMPLE', '\t'.join(str(part) for part in sample))
