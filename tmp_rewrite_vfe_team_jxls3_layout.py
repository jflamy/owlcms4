from copy import copy
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.cell.cell import MergedCell

TEMPLATE_PATH = r"C:\Dev\git\owlcms4\owlcms\src\main\resources\templates\teams\VFE_Teams-A4.xlsx"
MAX_COL = 27  # AA

wb = load_workbook(TEMPLATE_PATH)
ws = wb.active


def copy_row(src_row, dst_row):
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    for col in range(1, MAX_COL + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst.value = src.value
        dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        dst.comment = None if src.comment is None else Comment(src.comment.text, src.comment.author)


def clear_row(row_num):
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row_num, col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None
        cell.comment = None


# Rebuild the dynamic section to follow the Protocol_AllSessions pattern:
# one blank line before each gender title, no blank lines between athletes.
copy_row(12, 11)  # title row style
copy_row(14, 12)  # athlete row style and formulas
copy_row(23, 14)  # signature row

# Dynamic expressions
ws["A8"] = "${group.item.team}"
ws["A11"] = "${genderGroup.item.gender == 'F' ? t.get(\"Gender.Women\") : t.get(\"Gender.Men\")}"
ws["A12"] = "${l.gender}"
ws["B12"] = "${lIndex + 1}"
ws["C12"] = "${l.lastName}"
ws["D12"] = "${l.firstName}"
ws["F12"] = "${l.team}"
ws["G12"] = "${l.fullBirthDate}"
ws["H12"] = "${masters ? l.mastersLongRegistrationCategoryName : l.registrationCategory}"
ws["I12"] = None
ws["J12"] = "${l.entryTotal}"
ws["K12"] = None
ws["L12"] = "${l.teamAgeGroupsAsString}"
ws["M12"] = None

# JXLS3 comments
ws["A1"].value = None
ws["A1"].comment = Comment(
    'jx:area(lastCell="AA14")\n'
    'jx:each(items="lifters" var="group" groupBy="team" groupOrder="ASC" multisheet="group.item.team" lastCell="AA14")',
    "GitHub Copilot",
)
ws["A10"].value = None
ws["A10"].comment = Comment(
    'jx:each(items="group.items" var="genderGroup" groupBy="gender" groupOrder="ASC" lastCell="AA12")',
    "GitHub Copilot",
)
ws["A11"].comment = None
ws["A12"].comment = Comment(
    'jx:each(items="genderGroup.items" var="l" varIndex="lIndex" lastCell="AA12")',
    "GitHub Copilot",
)

# Remove the old legacy-conversion anchors and stray tag rows.
for ref in ["A16", "A17", "A24"]:
    ws[ref].value = None
    ws[ref].comment = None

# Move the signature merge up with the signature row.
if "A23:C23" in {str(rng) for rng in ws.merged_cells.ranges}:
    ws.unmerge_cells("A23:C23")
ws.merge_cells("A14:C14")

for row_num in range(15, 25):
    clear_row(row_num)

wb.save(TEMPLATE_PATH)
print(f"Rewrote {TEMPLATE_PATH}")
