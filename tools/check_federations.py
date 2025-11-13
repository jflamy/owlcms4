#!/usr/bin/env python3
"""
Check federation membership for rows in an Excel registration file using official membership lists.

Assumptions:
- Data starts on row 9 of the first sheet (Excel 1-indexed).
- Column E (Excel) contains a 3-letter IOC country code.
- Column P (Excel) contains one or more federation codes (comma/space/semicolon separated).

The script will:
- Load official federation membership lists from tools/federations/*.csv
- For each row starting at row 9, read column E (index 4) and column P (index 15).
- Determine which continental federation(s) the country should belong to based on membership lists.
- Report rows where the expected federation is not present in the federation column.

Usage:
    python tools/check_federations.py path/to/file.xlsx [--annotate out.xlsx]

Requirements:
    pip install pandas openpyxl
"""

import argparse
import sys
from pathlib import Path
import re

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas required. Install with: pip install pandas openpyxl")
    sys.exit(1)

SEPARATOR_RE = re.compile(r"[;,\s]+")

def load_federation_memberships(fed_dir: Path):
    """
    Load all federation membership CSVs from fed_dir.
    Returns dict: {IOC_code: set_of_federation_codes}
    """
    membership = {}
    fed_files = {
        'EWF': 'EWF_members.csv',
        'AWF': 'AWF_members.csv',
        'WFA': 'WFA_members.csv',
        'PAWF': 'PAWF_members.csv',
        'OWF': 'OWF_members.csv'
    }
    
    for fed_code, filename in fed_files.items():
        filepath = fed_dir / filename
        if not filepath.exists():
            print(f"Warning: {filepath} not found, skipping {fed_code}")
            continue
        df = pd.read_csv(filepath, dtype=str)
        for _, row in df.iterrows():
            ioc = str(row.iloc[0]).strip().upper()
            if ioc not in membership:
                membership[ioc] = set()
            membership[ioc].add(fed_code)
    
    return membership

def parse_feds_cell(cell):
    if pd.isna(cell):
        return set()
    s = str(cell).strip()
    if s == '':
        return set()
    parts = SEPARATOR_RE.split(s)
    return set([p.strip().upper() for p in parts if p.strip()])


def main():
    p = argparse.ArgumentParser()
    p.add_argument('workbook', type=Path, help='Excel file to check')
    p.add_argument('--annotate', '-a', type=Path, help='Optional output Excel to write annotated errors')
    p.add_argument('--fix', '-f', type=Path, help='Optional output Excel with corrected federation cells')
    args = p.parse_args()

    wb = args.workbook
    if not wb.exists():
        print(f"File not found: {wb}")
        sys.exit(2)

    # Load federation memberships from tools/federations/
    script_dir = Path(__file__).parent
    fed_dir = script_dir / 'federations'
    if not fed_dir.exists():
        print(f"Federation directory not found: {fed_dir}")
        sys.exit(2)
    
    membership = load_federation_memberships(fed_dir)
    print(f"Loaded membership for {len(membership)} countries across federations")

    # Read first sheet, skip first 8 rows so data begins at row 9
    df = pd.read_excel(wb, sheet_name=0, header=None, engine='openpyxl')

    start_idx = 8
    if start_idx >= len(df):
        print("No data starting at row 9")
        sys.exit(0)

    issues = []
    annotated = []
    for idx in range(start_idx, len(df)):
        row = df.iloc[idx]
        ioc = row.iloc[4] if len(row) > 4 else None
        fedcell = row.iloc[15] if len(row) > 15 else None
        ioc_str = str(ioc).strip().upper() if not pd.isna(ioc) else ''
        fedset = parse_feds_cell(fedcell)
        
        # Get expected federations from membership lists
        expected_feds = membership.get(ioc_str, set())
        
        ok = True
        note = ''
        missing = []
        wrong = []
        suggested = ''
        
        if not ioc_str or ioc_str == 'NAN':
            note = 'NO_COUNTRY'
            ok = True
        elif not expected_feds:
            note = f'UNKNOWN_IOC_{ioc_str}'
            ok = True  # Don't flag as error if we don't have membership data
        else:
            # Check if all expected federations are present in the cell
            missing = expected_feds - fedset
            # Check if there are continental federations that shouldn't be there
            continental_feds = {'EWF', 'AWF', 'WFA', 'PAWF', 'OWF'}
            wrong = (fedset & continental_feds) - expected_feds
            
            if missing or wrong:
                ok = False
                # Build suggested corrected string
                # Keep non-continental feds (like IWF), remove wrong ones, add missing ones
                corrected = (fedset - continental_feds - wrong) | expected_feds
                suggested = ','.join(sorted(corrected))
                
                parts = []
                if missing:
                    parts.append(f'MISSING_{",".join(sorted(missing))}')
                if wrong:
                    parts.append(f'WRONG_{",".join(sorted(wrong))}')
                note = ' | '.join(parts)
        
        if not ok:
            issues.append({
                'row': idx+1, 
                'ioc': ioc_str, 
                'feds_found': ','.join(sorted(fedset)), 
                'expected': ','.join(sorted(expected_feds)),
                'suggested': suggested,
                'missing': ','.join(sorted(missing)),
                'wrong': ','.join(sorted(wrong)),
                'note': note
            })
        
        annotated.append({
            'row': idx+1, 
            'ioc': ioc_str, 
            'feds_found': ','.join(sorted(fedset)), 
            'expected': ','.join(sorted(expected_feds)),
            'ok': ok, 
            'note': note
        })

    # Print summary
    print(f"\nScanned {len(df) - start_idx} rows (Excel row {start_idx+1}..{len(df)})")
    print(f"Issues found: {len(issues)}")
    if issues:
        print("\nAll issues:")
        for i in issues:
            print(f"  Excel row {i['row']:3}: IOC={i['ioc']:3} | Found: {i['feds_found']:20} | Should be: {i['suggested']:20} | {i['note']}")

    # Optionally write annotated Excel
    if args.annotate:
        out_df = pd.DataFrame(annotated)
        out_df.to_excel(args.annotate, index=False)
        print(f"\nWrote annotated output to {args.annotate}")
    
    # Optionally write corrected Excel
    if args.fix:
        # Load original workbook with openpyxl to preserve formatting
        from openpyxl import load_workbook
        wb_obj = load_workbook(wb)
        ws = wb_obj.worksheets[0]
        
        corrections_made = 0
        for issue in issues:
            row_num = issue['row']
            suggested = issue['suggested']
            # Column P is Excel column 16 (1-indexed)
            cell = ws.cell(row=row_num, column=16)
            old_value = cell.value
            cell.value = suggested
            corrections_made += 1
            print(f"  Row {row_num}: '{old_value}' -> '{suggested}'")
        
        wb_obj.save(args.fix)
        print(f"\nWrote corrected workbook with {corrections_made} fixes to {args.fix}")

if __name__ == '__main__':
    main()
