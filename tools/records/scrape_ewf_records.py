#!/usr/bin/env python3
"""
Scrape EWF (European Weightlifting Federation) records from the live EWF pages.

Usage:
    python scrape_ewf_records.py [--output <output.xlsx>] [--output-dir <dir>]

Behavior:
- Visits the six live EWF record pages for Senior/Junior/Youth and Women/Men
- Resolves the current "Download CSV" link from each page
- Parses the CSV rows into owlCMS record format
- Produces a single Excel workbook with one sheet per age group/gender combination
- Output defaults to I:/My Drive/records/EWF/EWF_scraped_YYYY-MM-DD_HHMMSS.xlsx on Windows
- On Linux, use --output-dir or OWLCMS_RECORDS_DIR to target a mounted Google Drive path for cron jobs

Requirements:
- openpyxl (pip install openpyxl)

Examples:
  python scrape_ewf_records.py
  python scrape_ewf_records.py --output "EWF_records.xlsx"
    python scrape_ewf_records.py --output-dir "/mnt/gdrive/records/EWF"
"""
from __future__ import annotations

import argparse
import csv
import html
import io
import os
import re
import shutil
import subprocess
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADER = [
    "Federation",
    "RecordName",
    "AgeGroup",
    "M/F",
    "ageLow",
    "ageUpper",
    "bwLow",
    "bwUpper",
    "Lift",
    "Record",
    "Name",
    "Born",
    "Nation",
    "Date",
    "Place",
    "Event",
    "Group",
]

AGE_GROUP_MAP = {
    "Senior": ("SR", 15, 999),
    "Junior": ("JR", 15, 20),
    "Youth": ("Youth", 13, 17),
}

AGE_GROUPS = ["Senior", "Junior", "Youth"]
GENDERS = ["Women", "Men"]
PAGE_URL_PATTERN = "https://ewf.sport/{age_group}-{gender}/"
WINDOWS_RECORDS_ROOT = Path("I:/My Drive/records")
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
    )
}
CSV_LINK_RE = re.compile(
    r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>\s*Download\s+CSV\s*</a>',
    re.IGNORECASE | re.DOTALL,
)
FALLBACK_CSV_RE = re.compile(r'https://[^"\']+output=csv[^"\']*', re.IGNORECASE)
CATEGORY_RE = re.compile(r'category\s*([+>]?\s*\d+|\d+\s*[+>])\s*kg', re.IGNORECASE)


def _download_text(url: str) -> str:
    request = urllib.request.Request(url, headers=REQUEST_HEADERS)
    with urllib.request.urlopen(request) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        return response.read().decode(charset, errors="replace")


def _normalize_date(date_str: str) -> str:
    """Normalize various date formats to yyyy-mm-dd."""
    if not date_str:
        return ""

    date_str = str(date_str).strip()
    if not date_str:
        return ""

    if " " in date_str:
        date_str = date_str.split(" ")[0]

    formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y/%m/%d",
        "%b %d, %Y",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return date_str


def _default_output_dir(federation: str) -> Path:
    configured_root = os.environ.get("OWLCMS_RECORDS_DIR")
    if configured_root:
        return Path(configured_root).expanduser() / federation

    if os.name == "nt":
        return WINDOWS_RECORDS_ROOT / federation

    return Path.home() / "records" / federation


def _clean_field(value: str) -> str:
    cleaned = str(value or "").strip()
    if cleaned.upper() in {"X", "-", "--", "---", "N/A"}:
        return ""
    return cleaned


def _parse_record_value(value: str) -> Optional[int]:
    cleaned = _clean_field(value).replace(",", ".")
    if not cleaned:
        return None

    match = re.search(r'\d+(?:\.\d+)?', cleaned)
    if not match:
        return None

    try:
        return int(float(match.group(0)))
    except ValueError:
        return None


def _extract_csv_url(page_url: str) -> str:
    page_html = _download_text(page_url)
    for match in CSV_LINK_RE.finditer(page_html):
        candidate = urljoin(page_url, html.unescape(match.group(1)).strip())
        if candidate.endswith("#") or candidate == page_url or "output=csv" not in candidate.lower():
            continue
        return candidate

    fallback = FALLBACK_CSV_RE.search(page_html)
    if fallback:
        return html.unescape(fallback.group(0))

    raise RuntimeError(f"Could not find Download CSV link on {page_url}")


def _fetch_csv_rows(csv_url: str) -> List[List[str]]:
    csv_text = _download_text(csv_url).lstrip("\ufeff")
    return list(csv.reader(io.StringIO(csv_text)))


def _detect_lift(cells: List[str]) -> Optional[str]:
    joined = " ".join(cell.lower() for cell in cells if cell)
    if "snatch" in joined:
        return "Snatch"
    if "clean" in joined or "c&j" in joined or "c+j" in joined:
        return "Clean & Jerk"
    if "total" in joined:
        return "Total"
    return None


def _is_header_row(cells: List[str]) -> bool:
    lowered = [cell.lower() for cell in cells if cell]
    return "record" in lowered and "name" in lowered


def _parse_category(text: str) -> Optional[Tuple[str, int]]:
    match = CATEGORY_RE.search(text)
    if not match:
        return None

    raw = re.sub(r'\s+', "", match.group(1))
    is_plus = raw.startswith(("+", ">")) or raw.endswith(("+", ">"))
    number = int(re.sub(r'[+>]', "", raw))
    return (f">{number}" if is_plus else str(number), number)


def _parse_csv_records(rows: List[List[str]], age_group_name: str, gender_name: str) -> List[Dict[str, Any]]:
    age_code, age_low, age_upper = AGE_GROUP_MAP[age_group_name]
    gender_code = "F" if gender_name == "Women" else "M"

    records: List[Dict[str, Any]] = []
    prev_bw_upper = 0
    current_bw_upper: Optional[str] = None
    current_bw_low = 0
    current_lift: Optional[str] = None
    awaiting_record_row = False

    for row in rows:
        cells = [str(cell).strip() for cell in row]
        nonempty = [cell for cell in cells if cell]
        if not nonempty:
            continue

        category = _parse_category(" ".join(nonempty))
        if category:
            current_bw_upper, bw_upper_num = category
            current_bw_low = prev_bw_upper
            prev_bw_upper = bw_upper_num
            current_lift = None
            awaiting_record_row = False
            print(f"  Category: {current_bw_upper} (bwLow={current_bw_low})")
            continue

        lift = _detect_lift(nonempty)
        if lift:
            current_lift = lift
            awaiting_record_row = False
            continue

        if current_bw_upper and current_lift and _is_header_row(nonempty):
            awaiting_record_row = True
            continue

        if not (current_bw_upper and current_lift and awaiting_record_row):
            continue

        trimmed = cells[:]
        while trimmed and not trimmed[0]:
            trimmed.pop(0)

        if not trimmed or _is_header_row([cell for cell in trimmed if cell]):
            continue

        record_val = _parse_record_value(trimmed[0])
        if record_val is None:
            continue

        name = _clean_field(trimmed[1] if len(trimmed) > 1 else "")
        born = _normalize_date(_clean_field(trimmed[2] if len(trimmed) > 2 else ""))
        nation = _clean_field(trimmed[3] if len(trimmed) > 3 else "")
        date = _normalize_date(_clean_field(trimmed[4] if len(trimmed) > 4 else ""))
        place = _clean_field(trimmed[5] if len(trimmed) > 5 else "")

        record = {
            "Federation": "EWF",
            "RecordName": "Europe",
            "AgeGroup": age_code,
            "M/F": gender_code,
            "ageLow": age_low,
            "ageUpper": age_upper,
            "bwLow": current_bw_low,
            "bwUpper": current_bw_upper,
            "Lift": current_lift,
            "Record": record_val,
            "Name": name,
            "Born": born,
            "Nation": nation,
            "Date": date,
            "Place": place,
            "Event": "",
            "Group": "",
        }
        records.append(record)
        print(f"    {current_bw_upper} {current_lift}: {record_val}kg by {name or 'Unknown'}")

        current_lift = None
        awaiting_record_row = False

    return records


def scrape_records() -> List[Dict[str, Any]]:
    """Scrape all live EWF record CSVs and convert them to owlCMS rows."""
    all_records: List[Dict[str, Any]] = []

    for age_group_name in AGE_GROUPS:
        for gender_name in GENDERS:
            print(f"\nProcessing {age_group_name} - {gender_name}...")
            page_url = PAGE_URL_PATTERN.format(age_group=age_group_name, gender=gender_name)

            try:
                csv_url = _extract_csv_url(page_url)
                print(f"  CSV: {csv_url}")
                rows = _fetch_csv_rows(csv_url)
                combo_records = _parse_csv_records(rows, age_group_name, gender_name)

                if not combo_records:
                    print(f"  No records found for {age_group_name} - {gender_name}")
                else:
                    print(f"  Parsed {len(combo_records)} records")
                    all_records.extend(combo_records)
            except Exception as exc:
                print(f"  Error processing {age_group_name} - {gender_name}: {exc}")

    return all_records


def write_to_excel(records: List[Dict[str, Any]], output_path: Path) -> None:
    """Write records to an owlCMS-style workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for record in records:
        key = (record["AgeGroup"], record["M/F"])
        grouped.setdefault(key, []).append(record)

    for (age_group, gender), group_records in grouped.items():
        sheet_name = f"{age_group} {'M' if gender == 'M' else 'W'}"
        ws = wb.create_sheet(title=sheet_name)

        for column_index, heading in enumerate(HEADER, start=1):
            ws.cell(row=1, column=column_index, value=heading)

        out_row = 2
        for record in group_records:
            for column_index, column_name in enumerate(HEADER, start=1):
                cell = ws.cell(row=out_row, column=column_index)
                value = record[column_name]

                if column_name in {"ageLow", "ageUpper", "bwLow", "Record"} and value != "":
                    cell.value = int(value)
                    cell.number_format = "0"
                else:
                    cell.value = value
            out_row += 1

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    print(f"\nWrote {output_path}")


def copy_to_destination(source_path: Path, destination_path: Path) -> Path:
    """Copy the scraped workbook to the requested destination path."""
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    print(f"Copied {source_path} to {destination_path}")
    return destination_path


def _maybe_open_output(path: Path) -> None:
    if os.name != "nt":
        return

    try:
        subprocess.run(["explorer", str(path)], check=False)
    except Exception as exc:
        print(f"Could not open file automatically: {exc}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Scrape live EWF records and write an owlCMS workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output owlCMS xlsx file (default: I:/My Drive/records/EWF/EWF_scraped_YYYY-MM-DD_HHMMSS.xlsx)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Directory for timestamped output, e.g. a mounted Google Drive path for cron jobs",
    )
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destination_xlsx: Optional[Path] = None
    output_dir = (args.output_dir.expanduser() if args.output_dir else _default_output_dir("EWF"))
    output_dir.mkdir(parents=True, exist_ok=True)
    if args.output:
        destination_xlsx = args.output.expanduser()
        output_xlsx = output_dir / f"{args.output.stem}_{timestamp}{args.output.suffix or '.xlsx'}"
    else:
        output_xlsx = output_dir / f"EWF_scraped_{timestamp}.xlsx"

    records = scrape_records()
    if not records:
        print("No records found!")
        return 1

    expected_total = len(AGE_GROUPS) * len(GENDERS) * 8 * 3
    print(f"\nTotal records scraped: {len(records)}")
    if len(records) != expected_total:
        print(f"WARNING: Expected {expected_total} records")

    write_to_excel(records, output_xlsx)

    final_path = output_xlsx
    if destination_xlsx:
        final_path = copy_to_destination(output_xlsx, destination_xlsx)

    _maybe_open_output(final_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())