#!/usr/bin/env python3
"""
Convert an AWF record-book PDF into an owlCMS-shaped Excel workbook.

Only the Senior, Junior, and Youth record pages are converted. Each output
sheet must contain 24 records: 8 bodyweight categories x 3 lifts.

Usage:
  python convert_awf_pdf_to_owlcms.py --input AWFRecordBook.pdf --output AWF_owlcms.xlsx

Requirements:
  pip install pdfplumber openpyxl
"""
from __future__ import annotations

import argparse
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Install with: pip install pdfplumber")
    raise SystemExit(1)


HEADER = [
    "Federation",
    "RecordName",
    "AgeGroup",
    "M/F",
    "ageLow",
    "ageUpper",
    "bwLow",
    "bwUpper",
    "Lift",
    "Record",
    "Name",
    "Born",
    "Nation",
    "Date",
    "Place",
    "Event",
    "Group",
]

AGE_GROUP_MAP = {
    "Senior": ("SR", 15, 999),
    "Junior": ("JR", 15, 20),
    "Youth": ("Youth", 13, 17),
}

LIFTS = ["Snatch", "C&J", "Total"]
HEADING_RE = re.compile(r"\bAWF\s+(FEMALE|MALE)\s+([A-Za-z ]+?)\s+Records\b", re.IGNORECASE)
CATEGORY_RE = re.compile(r"^[FM](\d+)(\+)?$", re.IGNORECASE)
RECORD_RE = re.compile(r"\d+(?:\.\d+)?")


def _clean_field(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_date(value: Any) -> str:
    text = _clean_field(value)
    if not text:
        return ""

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y/%m/%d",
        "%b %d, %Y",
        "%B %d, %Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return text


def _parse_record_value(value: Any) -> Optional[int]:
    match = RECORD_RE.search(_clean_field(value))
    if not match:
        return None

    return int(float(match.group(0)))


def _split_event_place(value: Any) -> Tuple[str, str]:
    text = _clean_field(value)
    if not text:
        return "", ""

    event, separator, place = text.partition(",")
    if not separator:
        return text, ""

    return event.strip(), place.strip()


def _find_page_context(text: str) -> Optional[Tuple[str, str, int, int, str]]:
    match = HEADING_RE.search(text)
    if not match:
        return None

    gender_text, age_group_text = match.groups()
    age_group_name = age_group_text.strip()
    if age_group_name not in AGE_GROUP_MAP:
        return None

    age_group_code, age_low, age_upper = AGE_GROUP_MAP[age_group_name]
    gender = "F" if gender_text.upper() == "FEMALE" else "M"
    return age_group_name, age_group_code, age_low, age_upper, gender


def _extract_categories(text: str) -> List[Tuple[int, str, int]]:
    categories: List[Tuple[int, str, int]] = []
    previous_upper = 0

    for line in text.splitlines():
        match = CATEGORY_RE.match(line.strip())
        if not match:
            continue

        upper_number = int(match.group(1))
        upper = f">{upper_number}" if match.group(2) else str(upper_number)
        categories.append((previous_upper, upper, upper_number))
        previous_upper = upper_number

    return categories


def _iter_record_tables(page: Any) -> Iterable[List[List[Any]]]:
    for table in page.extract_tables() or []:
        usable_rows = [row for row in table if row and any(_clean_field(cell) for cell in row)]
        if len(usable_rows) >= 3:
            yield usable_rows[:3]


def _normalize_name(name: Any) -> str:
    cleaned = _clean_field(name)
    if cleaned.lower() == "record standard":
        return "Australia Standard"
    return cleaned


def _normalize_record_date(name: str, record_date: Any, standard_date: str) -> str:
    if name == "Australia Standard":
        return standard_date
    return _normalize_date(record_date)


def parse_pdf(input_pdf: Path, federation: str, record_name: str, standard_date: str) -> Dict[str, List[Dict[str, Any]]]:
    records_by_sheet: Dict[str, List[Dict[str, Any]]] = {}

    with pdfplumber.open(input_pdf) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            context = _find_page_context(text)
            if not context:
                continue

            age_group_name, age_group_code, age_low, age_upper, gender = context
            sheet_name = f"{age_group_name} {'W' if gender == 'F' else 'M'}"
            categories = _extract_categories(text)
            tables = list(_iter_record_tables(page))

            if len(categories) != 8:
                print(f"WARNING: page {page_index} {sheet_name}: expected 8 categories, found {len(categories)}")
            if len(tables) != 8:
                print(f"WARNING: page {page_index} {sheet_name}: expected 8 record tables, found {len(tables)}")

            sheet_records: List[Dict[str, Any]] = []
            for category_index, (category, table) in enumerate(zip(categories, tables), start=1):
                bw_low, bw_upper, _ = category
                if len(table) != 3:
                    print(f"WARNING: page {page_index} {sheet_name} category {category_index}: expected 3 rows, found {len(table)}")

                for lift, row in zip(LIFTS, table):
                    cells = list(row) + ["", "", "", ""]
                    name = _normalize_name(cells[0])
                    record_value = _parse_record_value(cells[1])
                    if record_value is None:
                        print(f"WARNING: page {page_index} {sheet_name} {bw_upper} {lift}: missing record value")
                        continue

                    event, place = _split_event_place(cells[3])
                    sheet_records.append({
                        "Federation": federation,
                        "RecordName": record_name,
                        "AgeGroup": age_group_code,
                        "M/F": gender,
                        "ageLow": age_low,
                        "ageUpper": age_upper,
                        "bwLow": bw_low,
                        "bwUpper": bw_upper,
                        "Lift": lift,
                        "Record": record_value,
                        "Name": name,
                        "Born": "",
                        "Nation": "AUS",
                        "Date": _normalize_record_date(name, cells[2], standard_date),
                        "Place": place,
                        "Event": event,
                        "Group": "",
                    })

            records_by_sheet[sheet_name] = sheet_records
            print(f"Parsed {sheet_name}: {len(sheet_records)} records")

    return records_by_sheet


def write_workbook(records_by_sheet: Dict[str, List[Dict[str, Any]]], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet_order = ["Senior W", "Senior M", "Junior W", "Junior M", "Youth W", "Youth M"]
    for sheet_name in sheet_order:
        records = records_by_sheet.get(sheet_name, [])
        worksheet = workbook.create_sheet(title=sheet_name)

        for column_index, heading in enumerate(HEADER, start=1):
            worksheet.cell(row=1, column=column_index, value=heading)

        for row_index, record in enumerate(records, start=2):
            for column_index, column_name in enumerate(HEADER, start=1):
                value = record[column_name]
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                if column_name in {"ageLow", "ageUpper", "bwLow", "Record"} and value != "":
                    cell.value = int(value)
                    cell.number_format = "0"

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def validate_counts(records_by_sheet: Dict[str, List[Dict[str, Any]]]) -> bool:
    expected_sheets = ["Senior W", "Senior M", "Junior W", "Junior M", "Youth W", "Youth M"]
    ok = True

    for sheet_name in expected_sheets:
        count = len(records_by_sheet.get(sheet_name, []))
        if count != 24:
            print(f"ERROR: {sheet_name} has {count} records; expected 24")
            ok = False

    extra_sheets = sorted(set(records_by_sheet) - set(expected_sheets))
    if extra_sheets:
        print(f"ERROR: unexpected sheets parsed: {', '.join(extra_sheets)}")
        ok = False

    return ok


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert AWF record-book PDF pages for Senior, Junior, and Youth to owlCMS xlsx format.",
    )
    parser.add_argument("--input", type=Path, required=True, help="Input AWF record-book PDF")
    parser.add_argument("--output", type=Path, default=None, help="Output owlCMS xlsx file")
    parser.add_argument("--federation", default="AWF", help="Federation value for output rows")
    parser.add_argument("--record-name", default="Australia", help="RecordName value for output rows")
    parser.add_argument(
        "--standard-date",
        default="2026-05-06",
        help="Date to use for Australia Standard rows from Record Standard PDF rows",
    )
    args = parser.parse_args()

    input_pdf = args.input.expanduser()
    if not input_pdf.exists():
        print(f"ERROR: input PDF not found: {input_pdf}")
        return 1

    output_path = args.output.expanduser() if args.output else input_pdf.with_name(f"AWF_owlcms_{date.today():%Y-%m-%d}.xlsx")
    records_by_sheet = parse_pdf(input_pdf, args.federation, args.record_name, args.standard_date)
    if not validate_counts(records_by_sheet):
        return 1

    write_workbook(records_by_sheet, output_path)
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
