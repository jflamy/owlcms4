#!/usr/bin/env python
"""
Scrape IWF (International Weightlifting Federation) world records from their website.

Usage:
    python scrape_iwf_records.py [--output <output.xlsx>] [--output-dir <dir>]

Behavior:
- Scrapes https://iwf.sport/results/world-records/
- Iterates through all combinations of age groups and genders
- Extracts current world records for each combination
- Produces output in owlCMS Excel format matching the structure from convert_iwf_pdf_to_owlcms.py
- Output defaults to I:/My Drive/records/IWF/IWF_scraped_YYYY-MM-DD_HHMMSS.xlsx on Windows
- On Linux, use --output-dir or OWLCMS_RECORDS_DIR to target a mounted Google Drive path for cron jobs

Requirements:
- selenium (pip install selenium)
- openpyxl (pip install openpyxl)
- Chrome browser and chromedriver

Examples:
  python scrape_iwf_records.py
  python scrape_iwf_records.py --output "IWF_records.xlsx"
    python scrape_iwf_records.py --output-dir "/mnt/gdrive/records/IWF"
"""
from __future__ import annotations

import argparse
import os
import shutil
import subprocess
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
except ImportError:
    print("ERROR: selenium not installed. Install with: pip install selenium")
    exit(1)


HEADER = [
    "Federation",
    "RecordName",
    "AgeGroup",
    "M/F",
    "ageLow",
    "ageUpper",
    "bwLow",
    "bwUpper",
    "Lift",
    "Record",
    "Name",
    "Born",
    "Nation",
    "Date",
    "Place",
    "Event",
    "Group",
]

# Age group mappings
AGE_GROUP_MAP = {
    "Senior": ("SR", 15, 999),
    "Junior": ("JR", 15, 20),
    "Youth": ("Youth", 13, 17),
    "U15": ("U15", 0, 15),
}

WINDOWS_RECORDS_ROOT = Path("I:/My Drive/records")


def _normalize_date(date_str: str) -> str:
    """Normalize various date formats to yyyy-mm-dd."""
    if not date_str:
        return ""
    
    date_str = str(date_str).strip()
    if not date_str:
        return ""
    
    formats = [
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%b %d, %Y",
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return date_str


def _default_output_dir(federation: str) -> Path:
    configured_root = os.environ.get("OWLCMS_RECORDS_DIR")
    if configured_root:
        return Path(configured_root).expanduser() / federation

    if os.name == "nt":
        return WINDOWS_RECORDS_ROOT / federation

    return Path.home() / "records" / federation


def copy_to_destination(source_path: Path, destination_path: Path) -> Path:
    """Copy the scraped workbook to the requested destination path."""
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    print(f"Copied {source_path} to {destination_path}")
    return destination_path


def _maybe_open_output(path: Path) -> None:
    if os.name != "nt":
        return

    try:
        subprocess.run(["explorer", str(path)], check=False)
    except Exception as exc:
        print(f"Could not open file automatically: {exc}")


def scrape_records(url: str = "https://iwf.sport/results/world-records/") -> List[Dict[str, Any]]:
    """Scrape all world records from the IWF website."""
    
    # Setup Chrome driver
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # Run in background
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--log-level=3')  # Suppress selenium logs
    
    driver = webdriver.Chrome(options=options)
    all_records = []
    
    try:
        print(f"Loading {url}...")
        driver.get(url)
        
        # Wait for page to load
        wait = WebDriverWait(driver, 15)
        
        # Find the dropdowns
        # First dropdown: Record type (we want "Current")
        record_type_select = wait.until(
            EC.presence_of_element_located((By.ID, "ranking_curprog"))
        )
        record_type = Select(record_type_select)
        record_type.select_by_visible_text("Current")
        
        # Second dropdown: Age group
        age_group_select = driver.find_element(By.NAME, "ranking_agegroup")
        age_group = Select(age_group_select)
        age_group_options = [
            (opt.text.strip(), opt.get_attribute("value"))
            for opt in age_group.options
            if opt.text.strip()
        ]
        
        # Third dropdown: Gender
        gender_select = driver.find_element(By.ID, "ranking_gender")
        gender = Select(gender_select)
        gender_options = [
            (opt.text.strip(), opt.get_attribute("value"))
            for opt in gender.options
            if opt.text.strip()
        ]
        
        print(f"Age groups: {[name for name, _ in age_group_options]}")
        print(f"Genders: {[name for name, _ in gender_options]}")
        
        # Iterate through all combinations
        for age_group_name, age_group_value in age_group_options:
            for gender_name, gender_value in gender_options:
                print(f"\nProcessing {age_group_name} - {gender_name}...")

                query_url = f"{url}?{urlencode({'ranking_curprog': 'current', 'ranking_agegroup': age_group_value, 'ranking_gender': gender_value})}"
                driver.get(query_url)
                
                try:
                    # Load the server-rendered results page for the current combination.
                    wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.title__event h2"))
                    )

                    wait.until(
                        lambda d: any(
                            age_group_name.lower() in heading.text.strip().lower()
                            and gender_name.lower() in heading.text.strip().lower()
                            for heading in d.find_elements(By.CSS_SELECTOR, "div.title__event h2")
                        )
                    )

                    wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.results__title h2"))
                    )
                    
                    # Find all category headers (these contain bodyweight categories)
                    category_headers = driver.find_elements(By.CSS_SELECTOR, "div.results__title h2")
                    
                    # Parse age group info
                    age_code, age_low, age_upper = AGE_GROUP_MAP.get(age_group_name, (age_group_name, 0, 0))
                    
                    # Parse gender
                    gender_code = "M" if "Men" in gender_name or "Male" in gender_name else "F"
                    
                    # Filter to only category headers (contain "kg")
                    valid_headers = []
                    for header in category_headers:
                        header_text = header.text.strip()
                        if header_text and "kg" in header_text.lower() and "World Records" not in header_text:
                            valid_headers.append(header)
                    
                    # Get all card elements on the page
                    all_cards_raw = driver.find_elements(By.CSS_SELECTOR, "div.card")
                    
                    # Filter to only cards with title div (actual record cards)
                    all_cards = []
                    for card in all_cards_raw:
                        title_divs = card.find_elements(By.CSS_SELECTOR, "div.col-md-2.print__2.title")
                        if title_divs:
                            all_cards.append(card)
                    
                    print(f"  Found {len(valid_headers)} categories and {len(all_cards)} record cards total")
                    
                    prev_bw_num = 0
                    card_index = 0
                    
                    # Process each category header - assign 3 cards to each
                    for idx, header in enumerate(valid_headers):
                        header_text = header.text.strip()
                        
                        # Extract bodyweight category
                        import re
                        match = re.match(r'(\+?)(\d+)\s*kg', header_text, re.IGNORECASE)
                        if not match:
                            continue
                        
                        plus_sign = match.group(1)
                        bw_num = int(match.group(2))
                        
                        if plus_sign:
                            current_bw_upper = f">{bw_num}"
                        else:
                            current_bw_upper = str(bw_num)
                        
                        current_bw_low = prev_bw_num
                        prev_bw_num = bw_num
                        
                        print(f"  Category: {current_bw_upper} (bwLow={current_bw_low})")
                        
                        # Get the next 3 cards
                        cards_for_category = all_cards[card_index:card_index+3]
                        card_index += 3
                        
                        if len(cards_for_category) != 3:
                            print(f"    Warning: got {len(cards_for_category)} cards for category {current_bw_upper}, expected 3")
                        
                        for card in cards_for_category:
                                # Extract lift data from this card
                                title_divs = card.find_elements(By.CSS_SELECTOR, "div.col-md-2.print__2.title")
                                if not title_divs:
                                    continue
                                
                                title_div = title_divs[0]
                                paragraphs = title_div.find_elements(By.TAG_NAME, "p")
                                if len(paragraphs) < 2:
                                    continue
                                
                                lift_type = paragraphs[0].text.strip()
                                record_val_str = paragraphs[1].text.replace("Record:", "").replace("kg", "").strip()
                                
                                if not lift_type or not record_val_str:
                                    continue
                                
                                # Convert to integer (no decimals)
                                try:
                                    record_val = int(float(record_val_str))
                                except ValueError:
                                    print(f"    Warning: Could not parse record value: {record_val_str}")
                                    continue
                                
                                # Normalize lift type - handle both C&J and C&amp;J
                                if "Snatch" in lift_type:
                                    lift = "Snatch"
                                elif "C&J" in lift_type or "C&amp;J" in lift_type or "Clean" in lift_type or "Jerk" in lift_type:
                                    lift = "Clean & Jerk"
                                elif "Total" in lift_type:
                                    lift = "Total"
                                else:
                                    print(f"    Unknown lift type: '{lift_type}' - skipping")
                                    continue
                                
                                # Look for athlete name
                                athlete_divs = card.find_elements(By.CSS_SELECTOR, "div.col-md-3.print__3.not__cell__767")
                                athlete_name = ""
                                if athlete_divs:
                                    athlete_name = athlete_divs[0].text.strip()
                                
                                # Look for date and place (in col-md-7)
                                detail_divs = card.find_elements(By.CSS_SELECTOR, "div.col-md-7.print__7")
                                date_str = ""
                                place_str = ""
                                nation = ""
                                born_str = ""
                                
                                if detail_divs:
                                    detail_paragraphs = detail_divs[0].find_elements(By.TAG_NAME, "p")
                                    
                                    # Parse all paragraphs by content
                                    for para in detail_paragraphs:
                                        text = para.text.strip()
                                        
                                        if text.startswith("Event Place & Date:"):
                                            # Extract date and place (city)
                                            full_text = text.replace("Event Place & Date:", "").strip()
                                            
                                            # Format: "Jun 01, 2025 - World Standard" or "Jun 01, 2025 - CityName"
                                            if " - " in full_text:
                                                parts = full_text.split(" - ", 1)
                                                date_str = parts[0].strip()
                                                place_part = parts[1].strip() if len(parts) > 1 else ""
                                                # If it's "World Standard", leave place blank; otherwise it's a city
                                                if place_part != "World Standard":
                                                    place_str = place_part
                                            else:
                                                date_str = full_text
                                        
                                        elif text.startswith("Born:"):
                                            born_str = text.replace("Born:", "").strip()
                                        
                                        elif text.startswith("Nation:"):
                                            nation = text.replace("Nation:", "").strip()
                                
                                record = {
                                    "Federation": "IWF",
                                    "RecordName": "World",
                                    "AgeGroup": age_code,
                                    "M/F": gender_code,
                                    "ageLow": age_low,
                                    "ageUpper": age_upper,
                                    "bwLow": current_bw_low,
                                    "bwUpper": current_bw_upper,
                                    "Lift": lift,
                                    "Record": record_val,
                                    "Name": athlete_name,
                                    "Born": _normalize_date(born_str),
                                    "Nation": nation,
                                    "Date": _normalize_date(date_str),
                                    "Place": place_str,
                                    "Event": "",
                                    "Group": "",
                                }
                                all_records.append(record)
                                print(f"    {current_bw_upper} {lift}: {record_val}kg by {athlete_name}")
                
                except TimeoutException:
                    print(f"  No records found for {age_group_name} - {gender_name}")
                except Exception as e:
                    print(f"  Error processing {age_group_name} - {gender_name}: {e}")
    
    finally:
        driver.quit()
    
    return all_records


def write_to_excel(records: List[Dict[str, Any]], output_path: Path) -> None:
    """Write records to Excel file in owlCMS format."""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Group records by age group and gender
    grouped = {}
    for rec in records:
        key = (rec["AgeGroup"], rec["M/F"])
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(rec)
    
    # Create sheets
    for (age_group, gender), recs in grouped.items():
        sheet_name = f"{age_group} {'M' if gender == 'M' else 'W'}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Write header
        for c, h in enumerate(HEADER, start=1):
            ws.cell(row=1, column=c, value=h)
        
        # Write records
        out_row = 2
        for rec in recs:
            for cidx, col_name in enumerate(HEADER, start=1):
                cell = ws.cell(row=out_row, column=cidx)
                value = rec[col_name]
                
                # Ensure numeric columns are stored as numbers, not text
                if col_name in ("ageLow", "ageUpper", "Record") and value != "":
                    cell.value = int(value) if value else None
                    cell.number_format = '0'  # Integer format
                else:
                    cell.value = value
            out_row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"\nWrote {output_path}")


def main() -> int:
    p = argparse.ArgumentParser(
        description="Scrape IWF world records from their website.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--output", type=Path, default=None,
                   help="Output owlCMS xlsx file (default: I:/My Drive/records/IWF/IWF_scraped_YYYY-MM-DD_HHMMSS.xlsx)")
    p.add_argument("--output-dir", type=Path, default=None,
                   help="Directory for timestamped output, e.g. a mounted Google Drive path for cron jobs")
    args = p.parse_args()
    
    # Always use timestamp in filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destination_xlsx: Optional[Path] = None
    output_dir = (args.output_dir.expanduser() if args.output_dir else _default_output_dir("IWF"))
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.output:
        destination_xlsx = args.output.expanduser()
        output_xlsx = output_dir / f"{args.output.stem}_{timestamp}{args.output.suffix or '.xlsx'}"
    else:
        output_xlsx = output_dir / f"IWF_scraped_{timestamp}.xlsx"
    
    # Scrape records
    records = scrape_records()
    
    if not records:
        print("No records found!")
        return 1
    
    print(f"\nTotal records scraped: {len(records)}")
    
    # Write to Excel
    write_to_excel(records, output_xlsx)

    final_path = output_xlsx
    if destination_xlsx:
        final_path = copy_to_destination(output_xlsx, destination_xlsx)

    _maybe_open_output(final_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
