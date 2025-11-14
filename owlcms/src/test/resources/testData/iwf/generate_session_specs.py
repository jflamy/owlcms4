#!/usr/bin/env python3
"""
Generate a test Excel file with session specifications for NRegistrationFileProcessorTest
"""
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sessions"

# Row 1: Warning/info line (blank in this case)
ws['A1'] = "Session Specifications"
ws['A1'].font = Font(italic=True, size=9)

# Row 2: Headers
headers = [
    "sessionName", "platform", "description", "competitionTime", "weighinTime",
    "jury1", "jury2", "jury3", "jury4", "jury5",
    "ref1", "ref3",
    "reserve", "marshal", "marshal2", "timekeeper", 
    "techController", "techController2",
    "competitionSecretary", "competitionSecretary2",
    "doctor", "doctor2"
]

for col, header in enumerate(headers, start=1):
    ws.cell(row=2, column=col, value=header)

# Row 3+: Sample session data (16 sessions)
base_date = datetime(2025, 11, 15, 9, 0)

sessions = []
for i in range(1, 17):
    session_name = f"Session {i}"
    platform = "A" if i <= 8 else "B"
    description = f"Platform {platform} - Session {i}"
    comp_time = (base_date + timedelta(hours=i-1)).strftime("%Y-%m-%d %H:%M")
    weighin_time = (base_date + timedelta(hours=i-1, minutes=-60)).strftime("%Y-%m-%d %H:%M")
    
    row_data = [
        session_name,           # sessionName
        platform,               # platform
        description,            # description
        comp_time,              # competitionTime
        weighin_time,           # weighinTime
        f"Jury{i}_1",          # jury1
        f"Jury{i}_2",          # jury2
        f"Jury{i}_3",          # jury3
        f"Jury{i}_4",          # jury4
        f"Jury{i}_5",          # jury5
        f"Ref{i}_1",           # ref1
        f"Ref{i}_3",           # ref3
        f"Reserve{i}",         # reserve
        f"Marshal{i}",         # marshal
        f"Marshal{i}_2",       # marshal2
        f"TimeKeeper{i}",      # timekeeper
        f"TechCtrl{i}",        # techController
        f"TechCtrl{i}_2",      # techController2
        f"CompSec{i}",         # competitionSecretary
        f"CompSec{i}_2",       # competitionSecretary2
        f"Doctor{i}",          # doctor
        f"Doctor{i}_2",        # doctor2
    ]
    
    for col, value in enumerate(row_data, start=1):
        ws.cell(row=i+2, column=col, value=value)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save the file
output_path = "/c/Users/lamyj/git/owlcms4/owlcms/src/test/resources/testData/iwf/session_specs.xlsx"
wb.save(output_path)
print(f"Generated {output_path}")
print(f"  - Sheet: {ws.title}")
print(f"  - Row 1: Info line")
print(f"  - Row 2: {len(headers)} headers")
print(f"  - Rows 3-18: 16 sample sessions")
