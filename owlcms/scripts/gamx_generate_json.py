#!/usr/bin/env python3
"""
Generate GAMX parameter JSON files from the GAMX Excel workbooks.

Inputs (in owlcms/scripts/gamx-source/, NOT packaged into the jar):
- GAMX_calculator_allages_current.xlsx : TOTAL parameters (all variants)
- GAMX_CJ_Snatch.xlsx                  : SNATCH and C&J parameters

Outputs (default: owlcms/src/main/resources/gamx/, packaged into the jar):
- params-total-{variant}-{gender}.json  (4 variants × 2 genders = 8 files)
- params-snatch-{variant}-{gender}.json (2 variants × 2 genders = 4 files)
- params-cj-{variant}-{gender}.json     (2 variants × 2 genders = 4 files)

An alternate output directory may be passed as the first command-line argument,
e.g. for generating into a temp dir for comparison:
    python3 gamx_generate_json.py /tmp/gamx-new

Column mapping:
- TOTAL  : [bmass, mu, sigma, nu] (sen) or [age, bmass, mu, sigma, nu] (age/u17/mas)
- SNATCH : [bmass, mu, sigma, nu] (sen) or [age, bmass, mu, sigma, nu] (mas)
- CJ     : [bmass, mu, sigma, nu] (sen) or [age, bmass, mu, sigma, nu] (mas)

Sheet naming (TOTAL workbook): the JSON "variant" maps to the workbook tab as:
- sen -> params_sen_{gender}
- age -> params_iwf_{gender}   (the _iwf tabs ARE the age-adjusted 13+ tables)
- u17 -> params_U_{gender}
- mas -> params_mas_{gender}
"""

import json
import sys
from pathlib import Path
import openpyxl

def resource_dir():
    """Return the default gamx resources directory (JSON output, packaged into the jar)."""
    return Path(__file__).parent.parent / "src" / "main" / "resources" / "gamx"

def source_dir():
    """Return the gamx source-input directory (xlsx; NOT packaged into the jar)."""
    return Path(__file__).parent / "gamx-source"

def extract_sheet(workbook, sheet_name):
    """
    Extract numeric parameter rows from a worksheet.

    First row is the header; subsequent rows are numeric data rows.
    Rows with missing (None) cells, or with a column count that does not match
    the header, are skipped (handles trailing/blank rows in the workbook).

    Returns: list of lists (one list of floats per data row).
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]
    header = None
    data = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        numeric_row = []
        try:
            for v in row:
                if v is None:
                    break
                numeric_row.append(float(v))
            if numeric_row and len(numeric_row) == len(header):
                data.append(numeric_row)
        except (ValueError, TypeError):
            continue

    # Normalize row ordering so lookups (binary-search-by-age in both Java and JS)
    # always see a consistent layout, regardless of how the workbook tab is laid out:
    # - age tables [age, bmass, ...]  -> sorted age-major (age, then bmass)
    # - senior tables [bmass, ...]    -> sorted by bmass
    if data:
        ncol = len(data[0])
        if ncol >= 5:
            data.sort(key=lambda r: (r[0], r[1]))
        else:
            data.sort(key=lambda r: r[0])

    return data

def write_json_file(output_path, data):
    """Write array-of-arrays to JSON file."""
    with open(output_path, 'w') as f:
        json.dump(data, f, separators=(',', ':'))
    print(f"Wrote: {output_path}")

def main():
    if len(sys.argv) > 1:
        gamx_dir = Path(sys.argv[1])
    else:
        gamx_dir = resource_dir()
    gamx_dir.mkdir(parents=True, exist_ok=True)

    src_dir = source_dir()
    total_workbook_path = src_dir / "GAMX_calculator_allages_current.xlsx"
    snatch_cj_workbook_path = src_dir / "GAMX_CJ_Snatch.xlsx"

    if not total_workbook_path.exists():
        print(f"Error: TOTAL workbook not found: {total_workbook_path}", file=sys.stderr)
        return 1
    if not snatch_cj_workbook_path.exists():
        print(f"Error: SNATCH/CJ workbook not found: {snatch_cj_workbook_path}", file=sys.stderr)
        return 1

    # JSON variant -> TOTAL workbook tab prefix.
    # NOTE: the _iwf tabs are the age-adjusted (13+) tables ("age" variant).
    total_variants = {
        'sen': 'params_sen',
        'age': 'params_iwf',
        'u17': 'params_U',
        'mas': 'params_mas',
    }

    print("=== GENERATING TOTAL PARAMETERS (from GAMX_calculator_allages_current.xlsx) ===")
    total_wb = openpyxl.load_workbook(total_workbook_path, read_only=True)
    try:
        for variant, sheet_prefix in total_variants.items():
            for gender in ['men', 'wom']:
                sheet_name = f"{sheet_prefix}_{gender}"
                try:
                    data = extract_sheet(total_wb, sheet_name)
                except ValueError as e:
                    print(f"Error extracting {sheet_name}: {e}", file=sys.stderr)
                    return 1
                output_file = gamx_dir / f"params-total-{variant}-{gender}.json"
                write_json_file(output_file, data)
    finally:
        total_wb.close()

    snatch_cj_wb = openpyxl.load_workbook(snatch_cj_workbook_path, read_only=True)
    try:
        print("\n=== GENERATING SNATCH PARAMETERS (from GAMX_CJ_Snatch.xlsx) ===")
        for variant in ['sen', 'mas']:
            for gender in ['men', 'wom']:
                sheet_name = f"snatch_{variant}_{gender}"
                try:
                    data = extract_sheet(snatch_cj_wb, sheet_name)
                except ValueError as e:
                    print(f"Error extracting {sheet_name}: {e}", file=sys.stderr)
                    return 1
                output_file = gamx_dir / f"params-snatch-{variant}-{gender}.json"
                write_json_file(output_file, data)

        print("\n=== GENERATING C&J PARAMETERS (from GAMX_CJ_Snatch.xlsx) ===")
        for variant in ['sen', 'mas']:
            for gender in ['men', 'wom']:
                sheet_name = f"cj_{variant}_{gender}"
                try:
                    data = extract_sheet(snatch_cj_wb, sheet_name)
                except ValueError as e:
                    print(f"Error extracting {sheet_name}: {e}", file=sys.stderr)
                    return 1
                output_file = gamx_dir / f"params-cj-{variant}-{gender}.json"
                write_json_file(output_file, data)
    finally:
        snatch_cj_wb.close()

    print("\n=== SUMMARY ===")
    print(f"Output directory: {gamx_dir}")
    print("Generated 8 TOTAL files (sen, age, u17, mas × men, wom)")
    print("Generated 4 SNATCH files (sen, mas × men, wom)")
    print("Generated 4 CJ files (sen, mas × men, wom)")
    print("Total: 16 JSON files")
    return 0

if __name__ == '__main__':
    sys.exit(main())
