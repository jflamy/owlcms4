#!/usr/bin/env python3
"""
Converter for EWF-style records workbooks into owlCMS-shaped Excel workbooks.

Usage:
  python owlcms/scripts/records/convert_ewf_to_owlcms.py --input <file1> [<file2> ...] [--output <output.xlsx>]

Behavior:
- Accepts one or more EWF workbook files as input.
- Auto-detects gender (women/girls vs men/boys) from sheet names in each file.
- For each sheet, detects an AgeGroup (title) near the top and normalizes to JR/SR/U15/U23.
- Scans for weight-class rows (e.g., "49 KG"), then for lift sections: Snatch, C&J, Total.
- For each lift section, skips the block header row and takes the first data row with numeric record.
- Normalizes dates in Born and Date columns to yyyy-mm-dd format.
- Calculates bwLow: 0 for first weight category, previous bwUpper for others.
- Merges all inputs into a single output workbook with canonical owlCMS header structure.
- Output defaults to records/EWF_owlcms_YYYY-MM-DD.xlsx (or --output path if specified).
- If target file is locked, writes to -new.xlsx suffix instead.

Examples:
  # Convert both men's and women's files to dated output
  python owlcms/scripts/records/convert_ewf_to_owlcms.py --input records/EWF-RECORDS-W-300525-3.xlsx records/EWF-RECORDS-M-300525.xlsx

  # Specify custom output path
  python owlcms/scripts/records/convert_ewf_to_owlcms.py --input records/EWF-RECORDS-W-300525-3.xlsx records/EWF-RECORDS-M-300525.xlsx --output records/EWF_Europe_Records.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Optional, List

from openpyxl import load_workbook, Workbook
from datetime import datetime


HEADER = [
    "Federation",
    "RecordName",
    "AgeGroup",
    "M/F",
    "ageLow",
    "ageUpper",
    "bwLow",
    "bwUpper",
    "Lift",
    "Record",
    "Name",
    "Born",
    "Nation",
    "Date",
    "Place",
    "Event",
    "Group",
]


def _norm(x: Optional[str]) -> str:
    return "" if x is None else str(x).strip()


def _normalize_date(date_str: str) -> str:
    """Normalize various date formats to yyyy-mm-dd."""
    if not date_str:
        return ""
    
    date_str = str(date_str).strip()
    if not date_str:
        return ""
    
    # Remove time component if present (e.g., "2012-03-19 00:00:00")
    if " " in date_str:
        date_str = date_str.split(" ")[0]
    
    # Try common formats
    formats = [
        "%Y-%m-%d",      # 2025-04-30
        "%d.%m.%Y",      # 14.10.2006
        "%d/%m/%Y",      # 14/10/2006
        "%m/%d/%Y",      # 10/14/2006
        "%d-%m-%Y",      # 14-10-2006
        "%Y/%m/%d",      # 2006/10/14
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    # If nothing matched, return as-is
    return date_str


def _map_age(token: str) -> str:
    """Normalize age group tokens to Youth, U15, JR, U23, SR."""
    if not token:
        return ""
    t = token.lower()
    if "youth" in t or ("13" in t and "17" in t):
        return "Youth"
    if "u15" in t or "under 15" in t:
        return "U15"
    if "u23" in t or "under 23" in t:
        return "U23"
    if "senior" in t or "sr " in t:
        return "SR"
    if "junior" in t or "jr " in t:
        return "JR"
    if "boy" in t or "girl" in t:
        return "JR"
    if "men" in t or "women" in t:
        return "SR"
    return token


def _get_age_boundaries(agegroup: str) -> tuple[int, int]:
    """Return (ageLow, ageUpper) for each age group."""
    ag = agegroup.upper() if agegroup else ""
    if "YOUTH" in ag or "13" in ag or "17" in ag:
        return (13, 17)
    if "U15" in ag:
        return (0, 15)
    if "U23" in ag:
        return (15, 23)
    if "JR" in ag:
        return (15, 20)
    if "SR" in ag:
        return (15, 999)
    return (0, 0)


def _is_weight_row(cells: List[str]) -> Optional[str]:
    """Detect weight class row and normalize (e.g., '49 KG' -> '49', '>87' -> '>87')."""
    for cell in cells:
        if not cell:
            continue
        t = cell.lower().replace(" ", "")
        if "kg" in t or t.endswith("+") or t.startswith("+") or t.startswith(">"):
            cleaned = t.replace("kg", "").replace("+", "").replace(">", "").strip()
            try:
                n = int(float(cleaned))
                if "+" in cell or cell.startswith(">"):
                    return f">{n}"
                return f"{n}"
            except Exception:
                return cell
    return None


def _is_lift_row(cells: List[str]) -> Optional[str]:
    """Detect lift section header."""
    joined = " ".join(c.lower() for c in cells if c)
    if "snatch" in joined:
        return "Snatch"
    if "clean" in joined or "c&j" in joined or "c+j" in joined or "cj" in joined:
        return "Clean & Jerk"
    if "total" in joined:
        return "Total"
    return None


def _extract_agegroup(ws) -> str:
    """Extract and normalize age group from sheet."""
    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, min(8, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            s = str(v).strip()
            low = s.lower()
            if any(tok in low for tok in ("youth", "u15", "u23", "junior", "senior", "boy", "girl")) or (
                "u" in low and any(ch.isdigit() for ch in low)
            ):
                return _map_age(s)
    return ""


def _detect_gender_from_sheetname(sheetname: str) -> str:
    """Infer M/F from sheet name."""
    s = sheetname.lower()
    if any(tok in s for tok in ("women", "female", "girls", "g")):
        return "F"
    if any(tok in s for tok in ("men", "male", "boys", "b")):
        return "M"
    return ""


def convert_file(input_xlsx: Path, output_xlsx: Path) -> None:
    """Convert EWF workbook to owlCMS-shaped workbook."""
    wb_in = load_workbook(input_xlsx, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        out_ws = wb_out.create_sheet(title=sheet_name)

        # write header row
        for c, h in enumerate(HEADER, start=1):
            out_ws.cell(row=1, column=c, value=h)

        agegroup = _extract_agegroup(ws)
        sheet_gender = _detect_gender_from_sheetname(sheet_name)

        out_row = 2
        current_bw = None
        r = 1

        while r <= ws.max_row:
            cells = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

            # Detect weight row
            bw = _is_weight_row(cells)
            if bw:
                current_bw = bw
                r += 1
                continue

            # Detect lift section header
            lift = _is_lift_row(cells)
            if lift and current_bw:
                # Skip the block header row and find column positions
                header_row_idx = None
                header_tokens = ["name", "athlete", "record", "result", "nation", "country", "date", "place", "born", "yob"]
                lookahead_limit = min(ws.max_row, r + 8)

                for hr in range(r + 1, lookahead_limit + 1):
                    row_txt = [_norm(ws.cell(row=hr, column=c).value).lower() for c in range(1, ws.max_column + 1)]
                    if not any(row_txt):
                        continue
                    hits = sum(1 for t in header_tokens if any(t in cell for cell in row_txt if cell))
                    if hits >= 2:
                        header_row_idx = hr
                        break

                # Build column position map
                col_pos = {}
                if header_row_idx:
                    low = [_norm(ws.cell(row=header_row_idx, column=c).value).lower() for c in range(1, ws.max_column + 1)]

                    def find_col(keys):
                        for i, t in enumerate(low):
                            if not t:
                                continue
                            for k in keys:
                                if k in t:
                                    return i
                        return None

                    col_pos = {
                        "record": find_col(["record", "result", "kg"]),
                        "name": find_col(["name", "athlete"]),
                        "born": find_col(["born", "yob", "birth"]),
                        "nation": find_col(["nation", "country", "noc"]),
                        "date": find_col(["date"]),
                        "place": find_col(["place", "city", "location", "venue"]),
                    }

                # Find first data row with numeric record
                data_start = (header_row_idx + 1) if header_row_idx else (r + 1)
                rr = data_start
                found = False

                while rr <= ws.max_row and rr < data_start + 30:
                    data_cells = [_norm(ws.cell(row=rr, column=c).value) for c in range(1, ws.max_column + 1)]

                    # Skip empty rows
                    if not any(data_cells):
                        rr += 1
                        continue

                    # Skip repeated header rows
                    row_join = " ".join(val.lower() for val in data_cells if val)
                    if any(tok in row_join for tok in ("name", "athlete", "record", "result", "country", "nation")) and not any(
                        ch.isdigit() for ch in row_join
                    ):
                        rr += 1
                        continue

                    # Find record cell
                    rec_val = None
                    if col_pos.get("record") is not None and col_pos["record"] < len(data_cells):
                        rec_val = data_cells[col_pos["record"]]
                    else:
                        # Fallback: first cell with digit
                        for v in data_cells[:8]:
                            if v and any(ch.isdigit() for ch in v):
                                rec_val = v
                                break

                    if rec_val and any(ch.isdigit() for ch in rec_val):
                        # Extract other fields
                        def at_col(key, default_idx):
                            idx = col_pos.get(key)
                            if idx is not None and idx < len(data_cells):
                                return data_cells[idx]
                            if default_idx < len(data_cells):
                                return data_cells[default_idx]
                            return ""

                        rec = rec_val
                        name = at_col("name", 1)
                        born = at_col("born", 2)
                        nation = at_col("nation", 3)
                        date = at_col("date", 4)
                        place = at_col("place", 5)

                        row_vals = [
                            "EWF",
                            "Europe",
                            agegroup,
                            sheet_gender,
                            "",
                            "",
                            "",
                            current_bw,
                            lift,
                            rec,
                            name,
                            born,
                            nation,
                            date,
                            place,
                            "",
                            "",
                        ]
                        for cidx, val in enumerate(row_vals, start=1):
                            out_ws.cell(row=out_row, column=cidx, value=val)
                        out_row += 1
                        found = True
                        break

                    rr += 1

                r = (rr + 1) if found else (data_start + 1)
                continue

            r += 1

    wb_out.save(output_xlsx)


def merge_ewf_files(men_xlsx: Path, women_xlsx: Path, output_xlsx: Path) -> None:
    """Merge men's and women's EWF files into a single owlCMS workbook."""
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for input_file in [women_xlsx, men_xlsx]:
        if not input_file.exists():
            print(f"Warning: {input_file} not found, skipping")
            continue
        
        wb_in = load_workbook(input_file, data_only=True)
        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            out_ws = wb_out.create_sheet(title=sheet_name)
            
            # write header row
            for c, h in enumerate(HEADER, start=1):
                out_ws.cell(row=1, column=c, value=h)

            agegroup = _extract_agegroup(ws_in)
            sheet_gender = _detect_gender_from_sheetname(sheet_name)

            out_row = 2
            current_bw = None
            current_bw_low = 0
            prev_bw = None
            r = 1
            while r <= ws_in.max_row:
                cells = [_norm(ws_in.cell(row=r, column=c).value) for c in range(1, ws_in.max_column + 1)]

                bw = _is_weight_row(cells)
                if bw:
                    # Calculate bwLow: 0 for first, previous bwUpper for others
                    if prev_bw is None:
                        current_bw_low = 0
                    else:
                        # Extract number from previous bwUpper
                        prev_num_str = prev_bw.lstrip(">")
                        try:
                            current_bw_low = int(float(prev_num_str))
                        except Exception:
                            current_bw_low = 0
                    
                    current_bw = bw
                    prev_bw = bw
                    r += 1
                    continue

                lift = _is_lift_row(cells)
                if lift and current_bw:
                    # Skip the lift header row; r+1 is the column header row (Record, Name, Birth Date, Nation, Date, Place)
                    # r+2 is the first data row
                    rr = r + 2
                    found = False
                    while rr <= ws_in.max_row:
                        data_cells = [_norm(ws_in.cell(row=rr, column=c).value) for c in range(1, ws_in.max_column + 1)]
                        # Check if this row has any data (first data row after column header)
                        if any(data_cells):
                            # This is the first data row; extract fields
                            # Layout: Record | Name | Birth Date | Nation | Date | Place
                            rec = data_cells[0]
                            name = data_cells[1] if len(data_cells) > 1 else ""
                            born = _normalize_date(data_cells[2]) if len(data_cells) > 2 else ""
                            nation = data_cells[3] if len(data_cells) > 3 else ""
                            date = _normalize_date(data_cells[4]) if len(data_cells) > 4 else ""
                            place = data_cells[5] if len(data_cells) > 5 else ""
                            
                            # Override date for EWF Standard records
                            if name:
                                name_lower = str(name).strip().lower()
                                if "ewf" in name_lower and "standard" in name_lower:
                                    date = "2025-06-01"

                            agegroup = _extract_agegroup(ws_in)
                            age_low, age_upper = _get_age_boundaries(agegroup)

                            row_vals = [
                                "EWF",
                                "Europe",
                                agegroup,
                                sheet_gender,
                                age_low,
                                age_upper,
                                current_bw_low,
                                current_bw,
                                lift,
                                rec,
                                name,
                                born,
                                nation,
                                date,
                                place,
                                "",  # Event
                                "",  # Group
                            ]
                            for cidx, val in enumerate(row_vals, start=1):
                                out_ws.cell(row=out_row, column=cidx, value=val)
                            out_row += 1
                            found = True
                            break
                        rr += 1

                    r = (rr + 1) if found else (r + 1)
                    continue

                r += 1

    # Auto-fit column widths for all sheets
    from openpyxl.utils import get_column_letter
    for sheet in wb_out.sheetnames:
        ws = wb_out[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Add padding, cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb_out.save(output_xlsx)
    except PermissionError:
        # File is locked; save with -new suffix
        new_path = output_xlsx.with_stem(output_xlsx.stem + "-new")
        wb_out.save(new_path)
        print(f"Target is locked; wrote {new_path}")
        return


def main():
    from datetime import date
    
    p = argparse.ArgumentParser(
        description="Convert EWF-style records workbooks to owlCMS format.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert both men's and women's files
  %(prog)s --input records/EWF-RECORDS-W-300525-3.xlsx records/EWF-RECORDS-M-300525.xlsx

  # Specify custom output path
  %(prog)s --input records/EWF-RECORDS-W-300525-3.xlsx records/EWF-RECORDS-M-300525.xlsx --output records/EWF_Europe_Records.xlsx
        """
    )
    p.add_argument("--input", type=Path, nargs="+", required=True, 
                   help="Input EWF xlsx file(s) (one or more)")
    p.add_argument("--output", type=Path, default=None, 
                   help="Output owlCMS xlsx file (default: records/EWF_owlcms_YYYY-MM-DD.xlsx)")
    args = p.parse_args()

    input_files = args.input
    
    # Separate files by gender (detect from content)
    men_file = None
    women_file = None
    
    for f in input_files:
        wb = load_workbook(f, data_only=True)
        is_women = False
        for sheet_name in wb.sheetnames:
            if "women" in sheet_name.lower() or "girls" in sheet_name.lower():
                is_women = True
                break
        
        if is_women:
            women_file = f
        else:
            men_file = f
    
    if args.output:
        output_xlsx = args.output
    else:
        # Use the directory of the first input file as the output directory
        today = date.today().strftime("%Y-%m-%d")
        input_dir = input_files[0].parent
        output_xlsx = input_dir / f"EWF_owlcms_{today}.xlsx"

    merge_ewf_files(men_file, women_file, output_xlsx)
    print(f"Wrote {output_xlsx}")
    return 0


if __name__ == "__main__":
    main()
