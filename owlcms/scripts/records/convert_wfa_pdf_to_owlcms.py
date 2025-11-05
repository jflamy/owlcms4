#!/usr/bin/env python3
"""
Converter for WFA PDF records into owlCMS-shaped Excel workbooks.

Usage:
  python owlcms/scripts/records/convert_wfa_pdf_to_owlcms.py --input <pdf_file> [<pdf_file> ...] [--output <output.xlsx>]

Behavior:
- Accepts one or more WFA PDF files as input.
- Parses tables from PDFs to extract records.
- Auto-detects gender (Women vs Men) and age group (Senior/Junior/Youth) from filename and content.
- Normalizes age groups to JR/U23/Youth/SR based on context.
- Extracts bodyweight categories (including merged cells), lifts, and records.
- Produces output in owlCMS Excel format matching EWF_owlcms structure.
- Output defaults to I:/My Drive/records/WFA/WFA_owlcms_YYYY-MM-DD.xlsx (or --output path if specified).

Requirements:
- pdfplumber (pip install pdfplumber)

Examples:
  # Convert all WFA PDF files
  %(prog)s --input "I:/My Drive/records/WFA/1Men Senior*.pdf" "I:/My Drive/records/WFA/1Women Senior*.pdf"
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Optional, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Install with: pip install pdfplumber")
    exit(1)


HEADER = [
    "Federation",
    "RecordName",
    "AgeGroup",
    "M/F",
    "ageLow",
    "ageUpper",
    "bwLow",
    "bwUpper",
    "Lift",
    "Record",
    "Name",
    "Born",
    "Nation",
    "Date",
    "Place",
    "Event",
    "Group",
]


def _norm(x: Optional[str]) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s == "None" else s


def _detect_gender_from_filename(filename: str) -> str:
    """Infer M/F from filename."""
    f = filename.lower()
    if "women" in f:
        return "F"
    if "men" in f:
        return "M"
    return ""


def _detect_agegroup_from_filename(filename: str) -> str:
    """Infer age group from filename."""
    f = filename.lower()
    if "senior" in f:
        return "SR"
    if "junior" in f:
        return "JR"
    if "youth" in f:
        return "Youth"
    if "u15" in f:
        return "U15"
    if "u23" in f:
        return "U23"
    return ""


def _get_age_boundaries(agegroup: str) -> tuple[int, int]:
    """Return (ageLow, ageUpper) for each age group."""
    ag = agegroup.upper() if agegroup else ""
    if "YOUTH" in ag or "13" in ag or "17" in ag:
        return (13, 17)
    if "U15" in ag:
        return (0, 15)
    if "U23" in ag:
        return (15, 23)
    if "JR" in ag:
        return (15, 20)
    if "SR" in ag:
        return (15, 999)
    return (0, 0)


def _normalize_date(date_str: str) -> str:
    """Normalize various date formats to yyyy-mm-dd."""
    if not date_str:
        return ""
    
    date_str = str(date_str).strip()
    if not date_str:
        return ""
    
    # Remove time component if present
    if " " in date_str:
        date_str = date_str.split(" ")[0]
    
    formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return date_str


def convert_pdf(input_pdf: Path, wb_out: Workbook, federation: str = "WFA", record_name: str = "Africa") -> None:
    """Convert WFA PDF to owlCMS Excel workbook sheet."""
    gender = _detect_gender_from_filename(input_pdf.name)
    agegroup = _detect_agegroup_from_filename(input_pdf.name)
    age_low, age_upper = _get_age_boundaries(agegroup)
    
    sheet_name = f"{agegroup} {'M' if gender == 'M' else 'W'}"
    out_ws = wb_out.create_sheet(title=sheet_name)
    
    # Write header row
    for c, h in enumerate(HEADER, start=1):
        out_ws.cell(row=1, column=c, value=h)
    
    out_row = 2
    current_bw = None
    current_bw_low = 0
    prev_bw = None
    
    # Extract tables from PDF
    with pdfplumber.open(input_pdf) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if not tables:
                continue
            
            for table_idx, table in enumerate(tables):
                # Process each row in the table
                for row_idx, row in enumerate(table):
                    if not row or not any(row):
                        continue
                    
                    row_text = [_norm(cell) for cell in row]
                    
                    # Skip header row
                    if row_idx == 0:
                        continue
                    
                    # Check if this row has a new bodyweight category (column 0)
                    cat = row_text[0] if row_text else ""
                    has_new_bw = cat and ("kg" in cat.lower() or (cat and cat[0].isdigit()))
                    
                    if has_new_bw:
                        # This row has a bodyweight category
                        try:
                            cleaned = cat.lower().replace("kg", "").replace("+", "").replace(">", "").strip()
                            n = int(float(cleaned))
                            if "+" in cat or cat.startswith(">"):
                                bw = f">{n}"
                            else:
                                bw = f"{n}"
                            
                            # Calculate bwLow
                            if prev_bw is None:
                                current_bw_low = 0
                            else:
                                prev_num_str = prev_bw.lstrip(">")
                                try:
                                    current_bw_low = int(float(prev_num_str))
                                except Exception:
                                    current_bw_low = 0
                            
                            current_bw = bw
                            prev_bw = bw
                        except Exception:
                            pass
                    
                    # Extract lift and record data from any row with event/records
                    # Expected layout: [Category, Events, Records, Name, Nation, Event Date, Event Place]
                    event = row_text[1] if len(row_text) > 1 else ""
                    records = row_text[2] if len(row_text) > 2 else ""
                    
                    # Only process if we have a lift name and a record value
                    if event and records and current_bw:
                        name = row_text[3] if len(row_text) > 3 else ""
                        nation = row_text[4] if len(row_text) > 4 else ""
                        event_date = _normalize_date(row_text[5]) if len(row_text) > 5 else ""
                        event_place = row_text[6] if len(row_text) > 6 else ""
                        
                        # Parse lift type
                        event_lower = event.lower()
                        if "snatch" in event_lower:
                            lift = "Snatch"
                        elif "clean" in event_lower or "c&j" in event_lower or "cj" in event_lower:
                            lift = "Clean & Jerk"
                        elif "total" in event_lower:
                            lift = "Total"
                        else:
                            continue
                        
                        # Parse record value (remove "kg" if present)
                        rec_str = records.lower().replace("kg", "").strip() if records else ""
                        if not rec_str:
                            continue
                        
                        # Skip rows without athlete name
                        if not name:
                            continue
                        
                        # Use date 2025-06-01 for African Standard records
                        # and clear Place field
                        if "african standard" in name.lower():
                            event_date = "2025-06-01"
                            event_place = ""
                        
                        row_vals = [
                            federation,
                            record_name,
                            agegroup,
                            gender,
                            age_low,
                            age_upper,
                            current_bw_low,
                            current_bw,
                            lift,
                            rec_str,
                            name,
                            "",  # Born (not in PDF)
                            nation,
                            event_date,
                            event_place,
                            "",  # Event
                            "",  # Group
                        ]
                        
                        for cidx, val in enumerate(row_vals, start=1):
                            out_ws.cell(row=out_row, column=cidx, value=val)
                        out_row += 1
    
    # Auto-fit column widths
    for column in out_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        out_ws.column_dimensions[column_letter].width = adjusted_width


def main():
    from datetime import date
    
    p = argparse.ArgumentParser(
        description="Convert WFA PDF records to owlCMS format.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert all WFA PDF files
  %(prog)s --input "I:/My Drive/records/WFA/1Men Senior*.pdf" "I:/My Drive/records/WFA/1Women Senior*.pdf"
        """
    )
    p.add_argument("--input", type=Path, nargs="+", required=True,
                   help="Input WFA PDF file(s) (one or more)")
    p.add_argument("--output", type=Path, default=None,
                   help="Output owlCMS xlsx file (default: I:/My Drive/records/WFA/WFA_owlcms_YYYY-MM-DD.xlsx)")
    p.add_argument("--federation", type=str, default="WFA",
                   help="Federation name (default: WFA)")
    p.add_argument("--record-name", type=str, default="Africa",
                   help="Record name (default: Africa)")
    args = p.parse_args()

    input_files = args.input
    
    if args.output:
        output_xlsx = args.output
    else:
        today = date.today().strftime("%Y-%m-%d")
        input_dir = input_files[0].parent
        output_xlsx = input_dir / f"WFA_owlcms_{today}.xlsx"

    # Create workbook and process all PDF files
    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet
    
    processed_count = 0
    for input_file in input_files:
        if input_file.exists() and input_file.suffix.lower() == '.pdf':
            convert_pdf(input_file, wb_out, federation=args.federation, record_name=args.record_name)
            processed_count += 1
    
    if processed_count == 0:
        print("No PDF files found to process")
        return
    
    # Save the workbook
    try:
        wb_out.save(output_xlsx)
        print(f"Wrote {output_xlsx}")
    except PermissionError:
        new_path = output_xlsx.with_stem(output_xlsx.stem + "-new")
        wb_out.save(new_path)
        print(f"Target is locked; wrote {new_path}")


if __name__ == "__main__":
    main()
